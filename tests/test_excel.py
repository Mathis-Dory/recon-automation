import openpyxl
from recon import common


def test_write_enum_workbook(tmp_path):
    rows = [
        {"ip": "10.0.0.1", "port": 80, "state": "open",
         "http_title": "Login", "service": "http", "finding": ""},
        {"ip": "10.0.0.2", "port": 21, "state": "open",
         "http_title": "", "service": "ftp", "finding": "FTP ANON OK (listing: yes)"},
    ]
    out = str(tmp_path / "report.xlsx")
    result = common.write_enum_workbook(rows, out)
    assert result == out

    wb = openpyxl.load_workbook(out)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers == common.ENUM_COLUMNS
    assert ws.cell(row=2, column=1).value == "10.0.0.1"
    assert ws.cell(row=3, column=6).value == "FTP ANON OK (listing: yes)"
    # anon row highlighted, normal row not
    anon_fill = ws.cell(row=3, column=1).fill.fgColor.rgb
    normal_fill = ws.cell(row=2, column=1).fill.fgColor.rgb
    assert anon_fill != normal_fill


def test_empty_rows_just_header(tmp_path):
    out = str(tmp_path / "empty.xlsx")
    common.write_enum_workbook([], out)
    wb = openpyxl.load_workbook(out)
    assert [c.value for c in wb.active[1]] == common.ENUM_COLUMNS
