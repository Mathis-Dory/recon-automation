"""Shared helpers: target/port parsing, config, Excel/JSON, output, logging."""

import ipaddress
import os
import configparser
import logging
import shutil


def require_tools(names):
    """Raise RuntimeError if any external tool in `names` is not on PATH."""
    missing = [n for n in names if shutil.which(n) is None]
    if missing:
        raise RuntimeError("required tool(s) not found on PATH: " + ", ".join(missing))


def expand_range(spec):
    """Expand one target token into a list of IP strings."""
    spec = spec.strip()
    if "/" in spec:
        net = ipaddress.ip_network(spec, strict=False)
        return [str(h) for h in net.hosts()]
    if "-" in spec:
        left, right = spec.split("-", 1)
        left = left.strip()
        start = ipaddress.ip_address(left)
        if "." in right:
            end = ipaddress.ip_address(right.strip())
        else:
            prefix = left.rsplit(".", 1)[0]
            end = ipaddress.ip_address(f"{prefix}.{right.strip()}")
        if int(end) < int(start):
            raise ValueError(f"range end before start: {spec}")
        return [str(ipaddress.ip_address(i)) for i in range(int(start), int(end) + 1)]
    return [str(ipaddress.ip_address(spec))]


def parse_targets(range_=None, targets=None, infile=None):
    """Merge -r / -t / -iL inputs into a sorted, de-duplicated IP list."""
    tokens = []
    if range_:
        tokens.append(range_)
    if targets:
        tokens.extend(t for t in targets.split(",") if t.strip())
    if infile:
        with open(infile) as fh:
            for line in fh:
                line = line.strip()
                if line and not line.startswith("#"):
                    tokens.append(line)
    hosts = set()
    for tok in tokens:
        hosts.update(expand_range(tok))
    if not hosts:
        raise ValueError("no targets resolved from -r/-t/-iL")
    return sorted(hosts, key=lambda ip: ipaddress.ip_address(ip))


DEFAULT_WEB_PORTS = [
    80, 81, 88, 443, 446, 8080, 8081, 8082, 8083, 8085, 8443, 8888,
    9000, 9001, 9090, 8000, 8008, 8090, 8182, 8281, 7001, 10000,
]

SERVICE_PORTS = {
    "ftp": [21],
    "ssh": [22],
    "telnet": [23],
    "smb": [139, 445],
}


def parse_ports(spec):
    """Parse '80,443,8000-8002' into a sorted, de-duplicated int list."""
    ports = set()
    for tok in spec.split(","):
        tok = tok.strip()
        if not tok:
            continue
        if "-" in tok:
            a, b = tok.split("-", 1)
            lo, hi = int(a), int(b)
            rng = range(lo, hi + 1)
        else:
            rng = [int(tok)]
        for p in rng:
            if not 1 <= p <= 65535:
                raise ValueError(f"port out of range: {p}")
            ports.add(p)
    if not ports:
        raise ValueError("no ports parsed")
    return sorted(ports)


def default_enum_ports():
    """Sorted union of all service ports and the default web ports."""
    ports = set(DEFAULT_WEB_PORTS)
    for plist in SERVICE_PORTS.values():
        ports.update(plist)
    return sorted(ports)


DEFAULT_CONFIG_PATH = os.path.expanduser("~/.config/pentest-recon/config.ini")


def load_nessus_config(path=DEFAULT_CONFIG_PATH):
    """Load Nessus settings from an INI file."""
    if not os.path.exists(path):
        raise FileNotFoundError(f"config not found: {path}")
    parser = configparser.ConfigParser()
    parser.read(path)
    section = parser["nessus"] if parser.has_section("nessus") else {}
    cfg = {
        "url": section.get("url", "https://localhost:8834"),
        "access_key": section.get("access_key", ""),
        "secret_key": section.get("secret_key", ""),
        "template": section.get("template", "Basic Network Scan"),
    }
    if not cfg["access_key"] or not cfg["secret_key"]:
        raise ValueError("config missing access_key/secret_key in [nessus]")
    return cfg


DEFAULT_OUTPUT_ROOT = "~/tools/recon/output"
_OUTPUT_ROOT_ENV = "PT_RECON_OUTPUT"


def engagement_dir(name, root=None):
    """Create and return the output directory for an engagement.

    Precedence (highest first):
      1. ``root`` argument (typically from ``--outdir``)
      2. ``$PT_RECON_OUTPUT`` env var
      3. ``DEFAULT_OUTPUT_ROOT`` (~/tools/recon/output)
    """
    chosen = root or os.environ.get(_OUTPUT_ROOT_ENV) or DEFAULT_OUTPUT_ROOT
    path = os.path.join(os.path.expanduser(chosen), name)
    os.makedirs(path, exist_ok=True)
    return path


def get_logger(name):
    """Return a console logger with a single INFO handler."""
    logger = logging.getLogger(name)
    if not logger.handlers:
        handler = logging.StreamHandler()
        handler.setFormatter(logging.Formatter("[%(levelname)s] %(message)s"))
        logger.addHandler(handler)
        logger.setLevel(logging.INFO)
        logger.propagate = False
    return logger


import openpyxl
from openpyxl.styles import Font, PatternFill

ENUM_COLUMNS = ["ip", "port", "state", "http title", "service", "finding"]
_ROW_KEYS = ["ip", "port", "state", "http_title", "service", "finding"]
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_ANON_MARKERS = ("ANON", "NULL", "GUEST")


def write_enum_workbook(rows, out_path):
    """Write enumeration rows to an .xlsx report; return out_path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "results"
    ws.append(ENUM_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        values = [row.get(k, "") for k in _ROW_KEYS]
        ws.append(values)
        finding = str(row.get("finding", "")).upper()
        if any(m in finding for m in _ANON_MARKERS):
            for cell in ws[ws.max_row]:
                cell.fill = _RED_FILL
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{max(ws.max_row, 1)}"
    for col_idx, header in enumerate(ENUM_COLUMNS, start=1):
        width = max(len(header), 12)
        for row in rows:
            width = max(width, len(str(row.get(_ROW_KEYS[col_idx - 1], ""))))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(width + 2, 60)
    wb.save(out_path)
    return out_path
