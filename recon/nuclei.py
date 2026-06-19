"""nuclei target building and bootstrap."""
import os
import shutil
import subprocess

import openpyxl

from recon import common

_HTTPS_PORTS = {443, 8443}
_GO_INSTALL = "github.com/projectdiscovery/nuclei/v3/cmd/nuclei@latest"


def targets_from_enum(xlsx_path):
    """Build http(s)://ip:port URLs from web rows of an enum workbook."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    urls = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ip = row[idx["ip"]]
        port = row[idx["port"]]
        if port in common.DEFAULT_WEB_PORTS:
            scheme = "https" if port in _HTTPS_PORTS else "http"
            urls.append(f"{scheme}://{ip}:{port}")
    return urls


def ensure_nuclei(which=shutil.which, runner=subprocess.run):
    """Return the nuclei path, building it via go install if needed."""
    path = which("nuclei")
    if path:
        return path
    runner(["go", "install", _GO_INSTALL], check=True)
    gobin = os.path.expanduser("~/go/bin/nuclei")
    link = os.path.expanduser("~/tools/go-tools/bin/nuclei")
    try:
        os.makedirs(os.path.dirname(link), exist_ok=True)
        if not os.path.lexists(link):
            os.symlink(gobin, link)
    except OSError:
        pass
    return gobin


def build_nuclei_cmd(targets_file, out_path, severity="medium,high,critical", extra=None, nuclei_bin="nuclei"):
    """Construct the nuclei command line."""
    cmd = [
        nuclei_bin, "-l", targets_file,
        "-severity", severity,
        "-jsonl", "-o", out_path,
    ]
    if extra:
        cmd.extend(extra)
    return cmd
